import pandas as pd
from openpyxl import Workbook
import datetime
import numpy as np
import os
import difflib
import re
import time


def preencher_sheet(sheet, df, coluna_inicio, linha_inicio, limite_prenc_colunas):
    if sheet in book.sheetnames:
        ws1 = book[sheet]
    else:
        print('criando sheet')
        ws1 = book.create_sheet(sheet)
    #ws1.append(SAD_tabela_preco.columns.tolist())
    dados = df.values.tolist()
    for row_idx, row_data in enumerate(dados, start=linha_inicio):  # Começa na linha 2
        for col_idx, value in enumerate(row_data[:limite_prenc_colunas], start=coluna_inicio):  # Apenas as 15 primeiras colunas
            ws1.cell(row=row_idx, column=col_idx, value=value)  # Substitui os valores
    #return ws1

def formatar_moeda_brasileira(valor):
    valor_float = float(valor)
    return f'R$ {valor_float:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

def get_latest_file2(directory):
    files = [f for f in os.listdir(directory) 
             if f.endswith('.XLSX') and not f.startswith('~$') and f.startswith('Query')]
    if not files:
        raise FileNotFoundError(f"No Excel files found in {directory}")
    files = [os.path.join(directory, f) for f in files]
    latest_file = max(files, key=os.path.getmtime)
    print(f'Lendo {latest_file}')
    return latest_file

def inverter_data(data):
    if pd.isna(data):
        return np.nan
    
    if isinstance(data, str):
        if '/' in data:
            dia, mes, ano = data.split('/')
            return ano + mes + dia
        elif len(data) >= 10 and data[4] == '-' and data[7] == '-':
            return data[:10].replace('-', '')  # Converte 'yyyy-mm-dd' para 'yyyymmdd'
    
    return np.nan

data_atual = datetime.datetime.now()
data_formatada = data_atual.strftime("%d-%m-%Y")
data_hoje = pd.to_datetime(datetime.datetime.now().date()).strftime('%d/%m/%Y')
data_hoje = int(inverter_data(data_hoje))

ano_atual = data_atual.year
mes_atual = data_atual.strftime(f"%B_{ano_atual}")
mes_atual_pt = mes_atual.replace('January', 'Janeiro').replace('February', 'Fevereiro').replace('March', 'Março') \
    .replace('April', 'Abril').replace('May', 'Maio').replace('June', 'Junho').replace('July', 'Julho') \
    .replace('August', 'Agosto').replace('September', 'Setembro').replace('October', 'Outubro') \
    .replace('November', 'Novembro').replace('December', 'Dezembro')

caminho_arquivo = r'\\grupofleury\Diretoria Comercial\Equipe - Planejamento Comercial\06 - FERRAMENTAS\02 - Realinhamento\12 - Simulador\2025\Marca - OPA - Realinhamento 2025.xlsm'
caminho_arq = '\\\\grupofleury\\Auditoria\\Auditoria_Interna\\Auditoria_Continua\\Rogerio.Mello\\TRABALHOS REALIZADOS\\2 - [ADITIVOS] Leitura de PDFs e Imagens, Comparativos, Criação de Interface\\2.3 - Vef. exame disp\\BASES\\'
caminho_tabela_BI = '\\\\grupofleury\\Diretoria Comercial\\Projeto automação\\Extração Diária - Qlik\\' + data_formatada + '.xlsx'
#caminho_tabela_BI = '\\\\grupofleury\\Diretoria Comercial\\Equipe - Planejamento Comercial\\Projeto automação\\Extração Diária - Qlik\\11-02-2025.xlsx'
Query_sap_real_dir = fr'\\grupofleury\Dir_Exec_Suporte_Operacoes\GCC - Query SAP\2025\02 {mes_atual_pt}'
Query_sap_real_dir = fr'\\grupofleury\Dir_Exec_Suporte_Operacoes\GCC - Query SAP\2025\03 Março _2025'
Query_sap_real = get_latest_file2(Query_sap_real_dir)

###### LER DATAFRAMES
#tabela_bi = pd.read_excel(caminho_tabela_BI, keep_default_na=False, na_values=[])
tabela_bi = pd.read_excel(caminho_tabela_BI)
df_sap = pd.read_excel(Query_sap_real, usecols=['Nome da sigla de contrato', 'Descrição Convênio', 'Descrição Plano', 'Tabela de Preços', 'Validade do plano até', 'Válido até', 'Descrição Empresa', 'Capitation'], keep_default_na=False, na_values=[])
#df_sap = df_sap.rename(columns={'Tabela de preços ': 'Tabela de Preços'})
try:
    df_cubo = pd.read_parquet((caminho_arq + 'CUBO EXAMES.parquet'), )
except:
    df_cubo = pd.read_excel((caminho_arq + 'CUBO EXAMES.xlsx'), header = 6, keep_default_na=False, na_values=[])
    df_cubo.to_parquet((caminho_arq + 'CUBO EXAMES.parquet'), index=False)

###### TRATAMENTO QUERY_SAP
df_sap[['Validade do plano até', 'Válido até']] = (df_sap[['Validade do plano até', 'Válido até']].fillna('').astype(str))
df_sap['Validade do plano até'] = df_sap['Validade do plano até'].apply(inverter_data)
df_sap['Válido até'] = df_sap['Válido até'].apply(inverter_data)
df_sap[['Validade do plano até', 'Válido até']] = df_sap[['Validade do plano até', 'Válido até']].fillna(0).astype(int)
df_sap= df_sap[df_sap['Validade do plano até'] > data_hoje]
df_sap= df_sap[df_sap['Válido até'] > data_hoje]
df_sap = df_sap[df_sap['Capitation'] != 'S']
df_sap['Descrição Convênio'] = df_sap['Descrição Convênio'].str.strip()
df_sap['Descrição Convênio2'] = df_sap['Descrição Convênio'].str.lower()
nomes_coluna = df_sap['Descrição Convênio2'].astype(str).unique().tolist()

###### TRATAMENTO TABELA BI
tabela_bi['SIGLA_PRODUTO'] = tabela_bi['SIGLA_PRODUTO'].astype(str)
tabela_bi['CONVENIO'] = tabela_bi['CONVENIO'].str.strip()

###### CARREGAR PEDIDOS
pedidos = pd.read_csv(r'\\grupofleury\Auditoria\Auditoria_Interna\Auditoria_Continua\Rogerio.Mello\Demanda planejamento.del',encoding='latin-1' , sep='|', quotechar='"')
pedidos = pedidos.astype(str)
lista_pedidos = []
pedidos['Descrição_Convênio'] = pedidos['Descrição_Convênio'].str.strip().str.lower()
for index, row in pedidos.iterrows():
    if row['Descrição_Convênio'] != 'nan' and row['Opção'] == 'Volume':
        descricoes_convênio = row['Descrição_Convênio'].split('\\')
        nomes_proximos_pedido = []
        for descricao in descricoes_convênio:
            descricao = descricao.strip()
            nomes_parecidos = difflib.get_close_matches(descricao, nomes_coluna, n=3)
            print(nomes_parecidos)
            if nomes_parecidos:
                nome_escolhido = nomes_parecidos[0]
                filtro_pedido = df_sap[df_sap['Descrição Convênio2'] == nome_escolhido]
                nome_escolhido = filtro_pedido['Descrição Convênio'].iloc[0]
                nomes_proximos_pedido.append(nome_escolhido)
        if nomes_proximos_pedido:
            lista_pedidos.append(nomes_proximos_pedido)
    else:
        print("Nenhuma correspondência encontrada.")
        continue

###### PROCESSAR PEDIDOS
for string_convenio in lista_pedidos:
    print(string_convenio)
    parte_final = ' - '.join(string_convenio)
    parte_final_limpia = re.sub(r'[\\/[^a-zA-Z0-9\s-]]', '', parte_final)[:50]
    parte_final_limpia = re.sub(r'/', '', parte_final_limpia)
    parte_final_limpia = parte_final_limpia.strip()
    caminho_salvar_vol = rf'C:\Users\rogerio.mello\OneDrive - Grupo Fleury\Volumes\{data_formatada}\{parte_final_limpia}\{parte_final_limpia}.xlsx'
    caminho_salvar_vol_simulador = rf'C:\Users\rogerio.mello\OneDrive - Grupo Fleury\Volumes\{data_formatada}\{parte_final_limpia}'
    os.makedirs(caminho_salvar_vol_simulador, exist_ok=True)
    tabela_bi_conv = tabela_bi[tabela_bi['CONVENIO'].isin(string_convenio)]
    df_sap_conv = df_sap[df_sap['Descrição Convênio'].apply(lambda x: x.strip() if isinstance(x, str) else x).isin(string_convenio)]
    siglas = df_sap_conv['Nome da sigla de contrato'].unique()
    cep = df_sap_conv[["Descrição Convênio", "Descrição Empresa", "Descrição Plano", "Nome da sigla de contrato", "Tabela de Preços"]]
    lista_siglas = siglas.tolist()
    print(lista_siglas)
    #df_cubo2 = df_cubo[(df_cubo["CONVENIO"] == string_convenio) & (df_cubo['SIGLA CONTRATO'].isin(lista_siglas))]
    df_cubo2 = df_cubo[(df_cubo['SIGLA CONTRATO'].isin(lista_siglas))]
    print(df_cubo2)
    df_sap_conv = df_sap_conv.drop_duplicates(subset= 'Nome da sigla de contrato')
    df_cubo2_filtrado = pd.merge(
    df_cubo2,
    df_sap_conv[['Nome da sigla de contrato', 'Tabela de Preços']],
    left_on='SIGLA CONTRATO',  # Usando a coluna correta de df_cubo2
    right_on='Nome da sigla de contrato',  # Mantendo a coluna de df_sap
    how='right'
    )
    df_cubo2_filtrado['QTD TOTAL EXAMES'] = df_cubo2_filtrado['QTD TOTAL EXAMES'].replace('', 0).fillna(0)
    df_cubo2_filtrado['VL TOTAL EXAMES'] = df_cubo2_filtrado['VL TOTAL EXAMES'].replace('', 0).fillna(0)
    df_cubo2_filtrado['SIGLA EXAME'] = df_cubo2_filtrado['SIGLA EXAME'].fillna('').replace('', 'TOTAL@')
    df_cubo2_filtrado = df_cubo2_filtrado.groupby(['SIGLA EXAME', 'Tabela de Preços']).agg({
    'QTD TOTAL EXAMES': 'sum',                    # Somar VALOR 1
    'VL TOTAL EXAMES': 'sum'}).reset_index()
    total_por_tabela_preço = df_cubo2_filtrado.groupby('Tabela de Preços')[['VL TOTAL EXAMES', 'QTD TOTAL EXAMES']].sum().reset_index()
    total_por_tabela_preço['TOTAL VALOR TABELA'] = pd.to_numeric(total_por_tabela_preço['VL TOTAL EXAMES'])
    total_por_tabela_preço['TOTAL VOL TABELA'] = pd.to_numeric(total_por_tabela_preço['QTD TOTAL EXAMES'])

    indices_zero = total_por_tabela_preço[total_por_tabela_preço['Tabela de Preços'] == 0].index
    valores_originais = total_por_tabela_preço.loc[indices_zero, 'TOTAL VALOR TABELA'].copy()
    total_por_tabela_preço.loc[indices_zero, 'TOTAL VALOR TABELA'] = -1000
    total_por_tabela_preço['TABELA'] = total_por_tabela_preço['TOTAL VALOR TABELA'].rank(method='first', ascending=False).astype(int)
    total_por_tabela_preço.loc[indices_zero, 'TOTAL VALOR TABELA'] = valores_originais
    total_por_tabela_preço['TABELA'] = total_por_tabela_preço['TABELA'].apply(lambda x: f'TABELA {x}')
    total_por_tabela_preço['TOTAL VALOR TABELA'] = total_por_tabela_preço['TOTAL VALOR TABELA'].apply(formatar_moeda_brasileira)
    df_com_tabela = pd.merge(df_cubo2_filtrado, total_por_tabela_preço[['Tabela de Preços' ,'TOTAL VALOR TABELA', 'TABELA', 'TOTAL VOL TABELA']], on='Tabela de Preços', how='right')
    df_com_tabela = df_com_tabela[['SIGLA EXAME', 'QTD TOTAL EXAMES', 'Tabela de Preços', 'VL TOTAL EXAMES', 'TOTAL VALOR TABELA', 'TABELA', 'TOTAL VOL TABELA']]
    # Selecionar colunas necessárias
    df_resumo = df_com_tabela[['TOTAL VALOR TABELA', 'Tabela de Preços', 'TABELA']].drop_duplicates() #resumo pronto
    df_com_tabela = df_com_tabela[df_com_tabela['SIGLA EXAME'] != 'TOTAL@']
    print(df_resumo['TABELA'])
    df_resumo['TABELA_NUM'] = df_resumo['TABELA'].str.extract(r'(\d+)').astype(int)
    resumo_sorted = df_resumo.sort_values(by='TABELA_NUM')
    df_resumo = resumo_sorted.drop(columns='TABELA_NUM')
    print(df_resumo)
    df_resumo['data'] = pd.NaT  # Inicializa a coluna com valores NaT (Not a Time)
    df_resumo.at[0, 'data'] = datetime.datetime.today().date()  # Define a data de hoje na primeira linha
    cep = pd.merge(cep, df_resumo[['Tabela de Preços', 'TABELA']], on='Tabela de Preços', how= 'left')
    try:
        with pd.ExcelWriter(caminho_salvar_vol) as writer:
            df_resumo.to_excel(writer, sheet_name='RESUMO', index=False)
            df_com_tabela.to_excel(writer, sheet_name='TODOS', index=False)
            cep.to_excel(writer, sheet_name='CEP', index=False)
            tabelas_unicas = df_com_tabela['TABELA'].unique()
            for tabela in tabelas_unicas:
                df_tabela = df_com_tabela[df_com_tabela['TABELA'] == tabela]
                df_tabela.to_excel(writer, sheet_name=tabela, index=False)
            for tabela_preco in tabela_bi_conv['TABELA_PRECO'].unique():
                df_tabela_preco = tabela_bi_conv[tabela_bi_conv['TABELA_PRECO'] == tabela_preco]
                df_tabela_preco.to_excel(writer, sheet_name=str(tabela_preco), index=False)
    except Exception as e:
        print('Erro ao salvar arquivo')
        print(e)

    for tabela in df_resumo['Tabela de Preços'].unique():
        print(tabela)
        start_time = time.time()
        caminho_arquivo_r = fr'{caminho_salvar_vol_simulador}\{tabela}.xlsx'
        df_volume_tabela = df_com_tabela[df_com_tabela['Tabela de Preços'] == tabela]
        volume_colunas = ['SIGLA EXAME', 'QTD TOTAL EXAMES']
        SAD_tabela_preco = tabela_bi_conv[tabela_bi_conv['TABELA_PRECO'] == tabela] #SAD OK
        cep_tabela = cep[cep['Tabela de Preços'] == tabela]
        end_time = time.time()
        tempo_decorrido = end_time - start_time
        print(f"O tempo decorrido foi de {tempo_decorrido:.2f} segundos. PARA FILTROS")
        start_time = time.time()
        book = Workbook()
        book.active.title = "SAD"
        end_time = time.time()
        tempo_decorrido = end_time - start_time
        print(f"O tempo decorrido foi de {tempo_decorrido:.2f} segundos. PARA LOAD WORKBOOK")

        start_time = time.time()
        preencher_sheet('SAD', SAD_tabela_preco, 1, 2, 30) # a partir da primeira coluna, 2 linha (para não substituir o cabeçalho), com o limite de preencher até 30 colunas.
        preencher_sheet('CEP', cep_tabela, 1, 2 ,6)
        preencher_sheet('Volume', df_volume_tabela[volume_colunas], 2, 3, 3)
        preencher_sheet('Volume', df_volume_tabela[volume_colunas], 6, 3, 3)
        end_time = time.time()
        tempo_decorrido = end_time - start_time
        print(f"O tempo decorrido foi de {tempo_decorrido:.2f} segundos. PARA PREENCHER SHEETS")
        start_time = time.time()
        book.save(caminho_arquivo_r)
        end_time = time.time()
        tempo_decorrido = end_time - start_time
        print(f"O tempo decorrido foi de {tempo_decorrido:.2f} segundos. PARA SALVAR")